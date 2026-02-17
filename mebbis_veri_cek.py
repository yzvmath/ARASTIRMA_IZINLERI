# -*- coding: utf-8 -*-
"""
MEBBIS Araştırma İzni Veri Çekme Scripti
==========================================
Bu script MEBBIS sistemine giriş yapıp, araştırma izinleri
sayfasındaki bekleyen işlemleri ve detaylarını Excel'e aktarır.

Kullanım:
    python mebbis_veri_cek.py

Gereksinimler:
    pip install selenium openpyxl webdriver-manager
"""

import time
import os
import re
from datetime import datetime
import traceback

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
)
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ─── Ayarlar ──────────────────────────────────────────────────────────────────
TC_KIMLIK = "28585457194"
MEBBIS_URL = "https://mebbis.meb.gov.tr/ssologinBIDB.aspx?id=155"
HEDEF_URL = "https://arastirmaizinleri.meb.gov.tr/panel/arastirma-uygulamalari/bekleyen-islemler"
OUTPUT_DIR = r"d:\ARASTIMA_IZINLERI"
ZAMAN_DAMGASI = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"mebbis_verileri_{ZAMAN_DAMGASI}.xlsx")
# ─────────────────────────────────────────────────────────────────────────────

# Detay sayfasından çekilecek alanlar (sıralı)
DETAY_ALANLARI = [
    # ── Başvuru Bilgileri ──
    "Başvuru Numarası",
    "Başvuru Tarihi",
    "Başvuru Durumu",
    # ── Kişisel Bilgiler ──
    "TC Kimlik No",
    "Ad Soyad",
    "Telefon",
    "E-Posta",
    "Adres",
    # ── Başvuru Bilgileri (2) ──
    "Başvuru Şekli",
    "Başvurunun Yapıldığı Ülke",
    "Meslek",
    "Çalıştığı Kurum",
    # ── Araştırma Bilgileri ──
    "Araştırmanın Adı",
    "Eğitim Teknolojileri İle İlgili",
    "Araştırmanın Niteliği",
    "Akademik Başarı Ölçme",
    "Araştırmanın Konusu ve İlişkili Konular",
    "Anahtar Kelimeler",
    "Araştırmanın Yazım Dili",
    # ── Uygulama Bilgileri ──
    "Uygulama Yapılacak İl Sayısı",
    "Çalışma Grubu",
    "Teşkilat Türü",
    "Uygulama Yapılacak MEB Teşkilatı",
    "Uygulama Okul/Kurum Sayısı",
    "Özel Bilgiler",
    "Uygulama Süresi",
    # ── Belgeler ──
    "Araştırma Proje Bilgileri (Link)",
    "Veri Toplama Aracı (Link)",
    "Taahhütname (Link)",
    "Etik Kurul Onay (Link)",
    "Bilgilendirme ve Gönüllü Katılım Formu (Link)",
    "Veli Onam Formu (Link)",
    "Ölçek Kullanım İzni (Link)",
]


def tarayici_baslat():
    """Chrome tarayıcıyı başlatır."""
    print("\n🌐 Chrome tarayıcı başlatılıyor...")
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    chrome_options.add_experimental_option("detach", True)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver


def giris_yap(driver):
    """
    MEBBIS giriş sayfasını açar, TC kimlik numarasını otomatik doldurur
    ve kullanıcının oturumu açmasını bekler.
    """
    print(f"\n📌 MEBBIS giriş sayfası açılıyor: {MEBBIS_URL}")
    driver.get(MEBBIS_URL)
    time.sleep(3)

    # TC kimlik numarasını otomatik doldur
    try:
        tc_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "txtKullaniciAd"))
        )
        tc_input.clear()
        tc_input.send_keys(TC_KIMLIK)
        print(f"   ✅ TC Kimlik No otomatik girildi: {TC_KIMLIK}")
    except TimeoutException:
        print("   ⚠️ Kullanıcı adı alanı bulunamadı, manuel girin.")

    print("\n" + "=" * 60)
    print("  🔐 LÜTFEN OTURUMU AÇIN")
    print("     (Güvenlik kodu + Şifre + İki aşamalı doğrulama)")
    print("     TC Kimlik No otomatik girildi.")
    print("=" * 60)
    print("\n⏳ Oturumu açtıktan sonra buraya dönüp ENTER'a basın...")
    print()

    input("👉 Oturum açıldıysa ENTER'a basın: ")
    print("\n✅ Devam ediliyor...")
    time.sleep(1)


def hedef_sayfaya_git(driver):
    """Araştırma izinleri bekleyen işlemler sayfasına gider."""
    print(f"\n📌 Hedef sayfaya gidiliyor: {HEDEF_URL}")
    driver.get(HEDEF_URL)
    time.sleep(5)
    print(f"   Mevcut URL: {driver.current_url}")
    print(f"   Sayfa Başlığı: {driver.title}")


def ana_tablo_bul(driver):
    """Sayfadaki ana veri tablosunu bulur."""
    seciciler = [
        "table.table",
        "table.dataTable",
        "table[id*='Grid']",
        "table[id*='grid']",
        "div.table-responsive table",
        "table",
    ]
    for secici in seciciler:
        try:
            tablolar = driver.find_elements(By.CSS_SELECTOR, secici)
            for tablo in tablolar:
                satirlar = tablo.find_elements(By.TAG_NAME, "tr")
                if len(satirlar) >= 2:
                    tablo_id = tablo.get_attribute("id") or "isimsiz"
                    print(f"   ✅ Tablo bulundu (id: {tablo_id}, {len(satirlar)} satır)")
                    return tablo
        except Exception:
            pass
    print("   ❌ Uygun tablo bulunamadı!")
    return None


def tablo_basliklarini_oku(tablo):
    """Tablonun başlık satırını okur."""
    basliklar = []
    try:
        thead = tablo.find_elements(By.TAG_NAME, "thead")
        if thead:
            baslik_hucreleri = thead[0].find_elements(By.TAG_NAME, "th")
            if not baslik_hucreleri:
                baslik_hucreleri = thead[0].find_elements(By.TAG_NAME, "td")
        else:
            ilk_satir = tablo.find_elements(By.TAG_NAME, "tr")[0]
            baslik_hucreleri = ilk_satir.find_elements(By.TAG_NAME, "th")
            if not baslik_hucreleri:
                baslik_hucreleri = ilk_satir.find_elements(By.TAG_NAME, "td")

        for hucre in baslik_hucreleri:
            metin = hucre.text.strip()
            if metin:
                basliklar.append(metin)
            else:
                basliklar.append(f"Sütun_{len(basliklar) + 1}")
    except Exception as e:
        print(f"   ⚠️ Başlık okuma hatası: {e}")
    return basliklar


def tablo_satirlarini_oku(tablo):
    """Tablodaki veri satırlarını ve detay butonlarını okur."""
    satirlar_verisi = []
    try:
        tbody = tablo.find_elements(By.TAG_NAME, "tbody")
        if tbody:
            satirlar = tbody[0].find_elements(By.TAG_NAME, "tr")
        else:
            tum_satirlar = tablo.find_elements(By.TAG_NAME, "tr")
            satirlar = tum_satirlar[1:] if len(tum_satirlar) > 1 else []

        for satir in satirlar:
            hucreler = satir.find_elements(By.TAG_NAME, "td")
            if not hucreler:
                continue

            hucre_metinleri = []
            detay_butonu = None

            for hucre in hucreler:
                metin = hucre.text.strip()
                hucre_metinleri.append(metin)

                # Detay butonunu ara
                if not detay_butonu:
                    linkler = hucre.find_elements(By.TAG_NAME, "a")
                    butonlar = hucre.find_elements(By.TAG_NAME, "button")
                    tum_el = linkler + butonlar

                    for el in tum_el:
                        el_metin = (el.text or "").strip().lower()
                        el_title = (el.get_attribute("title") or "").lower()
                        el_class = (el.get_attribute("class") or "").lower()
                        el_href = (el.get_attribute("href") or "").lower()
                        el_onclick = (el.get_attribute("onclick") or "").lower()

                        if any(k in el_metin or k in el_title or k in el_class
                               for k in ["detay", "görüntüle", "incele", "göster",
                                          "detail", "view", "show", "eye", "search"]):
                            detay_butonu = el
                            break
                        elif any(k in el_href or k in el_onclick
                                 for k in ["detay", "detail", "basvuru-detay"]):
                            detay_butonu = el
                            break

                    # Son çare: satırdaki ilk anlamlı link
                    if not detay_butonu and linkler:
                        for link in linkler:
                            href = (link.get_attribute("href") or "")
                            if href and "#" not in href and "javascript:void" not in href:
                                detay_butonu = link
                                break

            satirlar_verisi.append({
                "hucre_metinleri": hucre_metinleri,
                "detay_butonu": detay_butonu,
            })
    except Exception as e:
        print(f"   ⚠️ Satır okuma hatası: {e}")
    return satirlar_verisi


def _linkleri_oku(parent_el, span_el):
    """
    Bir belge alanındaki linkleri [{"text": "...", "url": "..."}] listesi olarak döndürür.
    Önce parent_p içindeki <a> etiketlerini, bulamazsa ancestor div içindeki <a>'ları arar.
    """
    linkler_sonuc = []
    try:
        # Önce doğrudan parent'taki linkleri ara
        linkler = parent_el.find_elements(By.TAG_NAME, "a")
        if not linkler:
            # Daha geniş bir alanda ara
            parent_div = span_el.find_element(By.XPATH, "./ancestor::div[contains(@class,'mb-3')]")
            linkler = parent_div.find_elements(By.TAG_NAME, "a")
        for lnk in linkler:
            url = (lnk.get_attribute("href") or "").strip()
            metin = (lnk.text or "").strip()
            if url:
                if not metin:
                    metin = "Belge"
                linkler_sonuc.append({"text": metin, "url": url})
    except Exception:
        pass
    return linkler_sonuc if linkler_sonuc else ""


def detay_sayfasini_oku(driver):
    """
    Detay sayfasındaki tüm bilgileri yapısal olarak okur.
    Sayfa yapısı: <p class="ps-4"><span class="f-w-600">Etiket:</span> Değer</p>
    + Uygulama bilgileri tablosu (#ozetUygulamaBilgileriTable)
    + Belge linkleri (tıklanabilir hyperlink olarak saklanır)
    """
    detay = {}
    time.sleep(3)

    # ─────────────────────────────────────────────────────────────────────
    # 1) span.f-w-600 etiketlerinden anahtar-değer çiftlerini oku
    # ─────────────────────────────────────────────────────────────────────
    try:
        etiket_spanlari = driver.find_elements(By.CSS_SELECTOR, "span.f-w-600")
        for span in etiket_spanlari:
            try:
                etiket_raw = span.text.strip().rstrip(":").strip()
                if not etiket_raw or len(etiket_raw) > 120:
                    continue

                # Üst <p> veya <span> elementinden tam metni al
                parent_p = span.find_element(By.XPATH, "./..")
                tam_metin = parent_p.text.strip()

                # Etiket kısmını çıkart, kalanı değer
                if ":" in tam_metin:
                    idx = tam_metin.index(":")
                    deger = tam_metin[idx + 1:].strip()
                else:
                    deger = tam_metin.replace(etiket_raw, "").strip()

                if not deger:
                    continue

                # Etiketleri standart isimlere eşle
                etiket_lower = etiket_raw.lower()

                if "başvuru numarası" in etiket_lower:
                    detay["Başvuru Numarası"] = deger
                elif "başvuru tarihi" in etiket_lower:
                    detay["Başvuru Tarihi"] = deger
                elif "başvuru durumu" in etiket_lower:
                    detay["Başvuru Durumu"] = deger
                elif "tc kimlik" in etiket_lower:
                    detay["TC Kimlik No"] = deger
                elif "ad soyad" in etiket_lower:
                    detay["Ad Soyad"] = deger
                elif "telefon" in etiket_lower:
                    detay["Telefon"] = deger
                elif "e-posta" in etiket_lower or "eposta" in etiket_lower:
                    detay["E-Posta"] = deger
                elif "adres" in etiket_lower and "e-posta" not in etiket_lower:
                    detay["Adres"] = deger
                elif "başvuru şekli" in etiket_lower:
                    detay["Başvuru Şekli"] = deger
                elif "yapıldığı ülke" in etiket_lower:
                    detay["Başvurunun Yapıldığı Ülke"] = deger
                elif "meslek" in etiket_lower:
                    detay["Meslek"] = deger
                elif "çalıştığı kurum" in etiket_lower:
                    detay["Çalıştığı Kurum"] = deger
                elif "araştırmanın adı" in etiket_lower:
                    detay["Araştırmanın Adı"] = deger
                elif "eğitim teknolojileri" in etiket_lower:
                    detay["Eğitim Teknolojileri İle İlgili"] = deger
                elif "araştırmanın niteliği" in etiket_lower:
                    detay["Araştırmanın Niteliği"] = deger
                elif "akademik başarı" in etiket_lower:
                    detay["Akademik Başarı Ölçme"] = deger
                elif "konusu ve ilişkili" in etiket_lower:
                    # Konu birden fazla satır olabilir, temizle
                    deger_temiz = " | ".join([s.strip() for s in deger.split("\n") if s.strip()])
                    detay["Araştırmanın Konusu ve İlişkili Konular"] = deger_temiz
                elif "anahtar kelime" in etiket_lower:
                    detay["Anahtar Kelimeler"] = deger
                elif "yazım dili" in etiket_lower:
                    detay["Araştırmanın Yazım Dili"] = deger
                elif "il sayısı" in etiket_lower:
                    detay["Uygulama Yapılacak İl Sayısı"] = deger
                elif "araştırma proje" in etiket_lower:
                    detay["Araştırma Proje Bilgileri (Link)"] = _linkleri_oku(parent_p, span)
                elif "veri toplama aracı" in etiket_lower or "veri toplama araç" in etiket_lower:
                    detay["Veri Toplama Aracı (Link)"] = _linkleri_oku(parent_p, span)
                elif "taahhütname" in etiket_lower:
                    detay["Taahhütname (Link)"] = _linkleri_oku(parent_p, span)
                elif "etik kurul" in etiket_lower:
                    detay["Etik Kurul Onay (Link)"] = _linkleri_oku(parent_p, span)
                elif "bilgilendirme" in etiket_lower and "gönüllü" in etiket_lower:
                    detay["Bilgilendirme ve Gönüllü Katılım Formu (Link)"] = _linkleri_oku(parent_p, span)
                elif "veli onam" in etiket_lower:
                    detay["Veli Onam Formu (Link)"] = _linkleri_oku(parent_p, span)
                elif "alanyazın" in etiket_lower or "kullanıma ilişkin izin" in etiket_lower:
                    detay["Ölçek Kullanım İzni (Link)"] = _linkleri_oku(parent_p, span)

            except (StaleElementReferenceException, NoSuchElementException):
                continue
            except Exception:
                continue
    except Exception as e:
        print(f"   ⚠️ Etiket okuma hatası: {e}")

    # ─────────────────────────────────────────────────────────────────────
    # 2) Uygulama bilgileri tablosunu oku (#ozetUygulamaBilgileriTable)
    # ─────────────────────────────────────────────────────────────────────
    try:
        uygulama_tablo = None
        # Önce ID ile ara
        try:
            uygulama_tablo = driver.find_element(By.ID, "ozetUygulamaBilgileriTable")
        except NoSuchElementException:
            # Sayfadaki tabloları ara
            tablolar = driver.find_elements(By.CSS_SELECTOR, "table.table-bordered, table.table-sm")
            for t in tablolar:
                text = t.text.lower()
                if "çalışma grubu" in text or "teşkilat" in text:
                    uygulama_tablo = t
                    break

        if uygulama_tablo:
            tbody = uygulama_tablo.find_elements(By.TAG_NAME, "tbody")
            if tbody:
                satirlar = tbody[0].find_elements(By.TAG_NAME, "tr")
            else:
                satirlar = uygulama_tablo.find_elements(By.TAG_NAME, "tr")[1:]  # başlığı atla

            # Birden fazla uygulama satırı olabilir, hepsini birleştir
            calisma_gruplari = []
            teskilat_turleri = []
            meb_teskilatlari = []
            sayilar = []
            ozel_bilgiler = []
            uygulama_sureleri = []

            for satir in satirlar:
                hucreler = satir.find_elements(By.TAG_NAME, "td")
                if len(hucreler) >= 6:
                    calisma_gruplari.append(hucreler[0].text.strip())
                    teskilat_turleri.append(hucreler[1].text.strip())
                    meb_teskilatlari.append(hucreler[2].text.strip())
                    sayilar.append(hucreler[3].text.strip())
                    ozel_bilgiler.append(hucreler[4].text.strip())
                    uygulama_sureleri.append(hucreler[5].text.strip())

            detay["Çalışma Grubu"] = " | ".join(calisma_gruplari) if calisma_gruplari else ""
            detay["Teşkilat Türü"] = " | ".join(teskilat_turleri) if teskilat_turleri else ""
            detay["Uygulama Yapılacak MEB Teşkilatı"] = " | ".join(meb_teskilatlari) if meb_teskilatlari else ""
            detay["Uygulama Okul/Kurum Sayısı"] = " | ".join(sayilar) if sayilar else ""
            detay["Özel Bilgiler"] = " | ".join(ozel_bilgiler) if ozel_bilgiler else ""
            detay["Uygulama Süresi"] = " | ".join(uygulama_sureleri) if uygulama_sureleri else ""
    except Exception as e:
        print(f"   ⚠️ Uygulama tablosu okuma hatası: {e}")

    return detay


def sayfalama_kontrol(driver):
    """Sayfalama varsa sonraki sayfaya geçer."""
    seciciler = [
        "a.page-link",
        "li.page-item a",
        "a[aria-label='Next']",
        ".pagination a",
        "a.next",
        "a.paginate_button.next",
    ]
    for secici in seciciler:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, secici)
            for el in elements:
                metin = el.text.strip().lower()
                aria = (el.get_attribute("aria-label") or "").lower()
                if any(k in metin or k in aria for k in ["next", "sonraki", "ileri", "›", "»", ">"]):
                    parent = el.find_element(By.XPATH, "..")
                    parent_class = (parent.get_attribute("class") or "").lower()
                    if "disabled" not in parent_class:
                        print("   📄 Sonraki sayfaya geçiliyor...")
                        el.click()
                        time.sleep(4)
                        return True
        except Exception:
            pass
    return False


def _hucre_yaz(ws, row, col, deger, veri_font, link_font, veri_alignment, ince_border):
    """
    Bir hücreye değer yazar. Değer link listesi ise tıklanabilir hyperlink yapar.
    - Tek link: HYPERLINK formülü ile tıklanabilir metin
    - Çoklu link: Her biri ayrı satırda HYPERLINK formülü (ilk linke hyperlink)
    - Normal metin: Düz metin olarak yazar
    """
    hucre = ws.cell(row=row, column=col)
    hucre.alignment = veri_alignment
    hucre.border = ince_border

    if isinstance(deger, list) and deger and isinstance(deger[0], dict):
        # Link listesi
        if len(deger) == 1:
            # Tek link → tıklanabilir hyperlink
            link = deger[0]
            hucre.value = f'=HYPERLINK("{link["url"]}", "{link["text"].replace(chr(34), chr(39))}")'  # " → '
            hucre.font = link_font
        else:
            # Çoklu link → ilk linke hyperlink, diğerlerini alt satırlara yaz
            ilk = deger[0]
            hucre.hyperlink = ilk["url"]
            hucre.value = "\n".join(lnk["text"] for lnk in deger)
            hucre.font = link_font
    else:
        hucre.value = deger if deger else ""
        hucre.font = veri_font


def excel_olustur(basliklar, tum_satirlar):
    """
    Tüm verileri tek satır halinde Excel'e yazar.
    Ana tablo sütunları + detay sütunları yan yana.
    Link alanları tıklanabilir hyperlink olarak yazılır.
    """
    print(f"\n📝 Excel dosyası oluşturuluyor: {OUTPUT_FILE}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bekleyen İşlemler"

    # ─── Stiller ─────────────────────────────────────────────────────────
    baslik_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    baslik_fill_ana = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    baslik_fill_kisisel = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    baslik_fill_basvuru = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    baslik_fill_arastirma = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    baslik_fill_uygulama = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    baslik_fill_belge = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    baslik_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    veri_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    veri_alignment = Alignment(vertical="center", wrap_text=True)
    ince_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ─── Renk haritası (detay alan adına göre) ───────────────────────────
    def detay_renk(alan):
        if alan in ["Başvuru Numarası", "Başvuru Tarihi", "Başvuru Durumu"]:
            return baslik_fill_ana
        elif alan in ["TC Kimlik No", "Ad Soyad", "Telefon", "E-Posta", "Adres"]:
            return baslik_fill_kisisel
        elif alan in ["Başvuru Şekli", "Başvurunun Yapıldığı Ülke", "Meslek", "Çalıştığı Kurum"]:
            return baslik_fill_basvuru
        elif alan in ["Araştırmanın Adı", "Eğitim Teknolojileri İle İlgili",
                       "Araştırmanın Niteliği", "Akademik Başarı Ölçme",
                       "Araştırmanın Konusu ve İlişkili Konular",
                       "Anahtar Kelimeler", "Araştırmanın Yazım Dili"]:
            return baslik_fill_arastirma
        elif alan in ["Uygulama Yapılacak İl Sayısı", "Çalışma Grubu", "Teşkilat Türü",
                       "Uygulama Yapılacak MEB Teşkilatı", "Uygulama Okul/Kurum Sayısı",
                       "Özel Bilgiler", "Uygulama Süresi"]:
            return baslik_fill_uygulama
        else:
            return baslik_fill_belge

    # ─── Başlıklar: Ana tablo + Detay alanları ──────────────────────────
    ana_basliklar = list(basliklar)
    tum_basliklar = list(ana_basliklar) + list(DETAY_ALANLARI)

    # Ekstra detay alanları (DETAY_ALANLARI'nda olmayan ama veride bulunan)
    ekstra_alanlar = []
    for satir in tum_satirlar:
        for key in satir.get("detay_verileri", {}):
            if key not in DETAY_ALANLARI and key not in ekstra_alanlar:
                ekstra_alanlar.append(key)
    tum_basliklar.extend(ekstra_alanlar)

    # Başlık satırını yaz
    for col, baslik in enumerate(tum_basliklar, 1):
        hucre = ws.cell(row=1, column=col, value=baslik)
        hucre.font = baslik_font
        hucre.alignment = baslik_alignment
        hucre.border = ince_border

        if col <= len(ana_basliklar):
            hucre.fill = baslik_fill_ana
        else:
            detay_alan = baslik
            hucre.fill = detay_renk(detay_alan)

    # ─── Verileri yaz ────────────────────────────────────────────────────
    for row_idx, satir in enumerate(tum_satirlar, 2):
        hucre_verileri = satir["hucre_metinleri"]
        detay_verileri = satir.get("detay_verileri", {})

        # Ana tablo sütunları
        for col, deger in enumerate(hucre_verileri, 1):
            if col <= len(ana_basliklar):
                hucre = ws.cell(row=row_idx, column=col, value=deger)
                hucre.font = veri_font
                hucre.alignment = veri_alignment
                hucre.border = ince_border

        # Detay sütunları (sabit alanlar)
        offset = len(ana_basliklar)
        for d_col, d_alan in enumerate(DETAY_ALANLARI):
            deger = detay_verileri.get(d_alan, "")
            col_num = offset + d_col + 1
            _hucre_yaz(ws, row_idx, col_num, deger, veri_font, link_font, veri_alignment, ince_border)

        # Ekstra detay alanları
        offset2 = offset + len(DETAY_ALANLARI)
        for e_col, e_alan in enumerate(ekstra_alanlar):
            deger = detay_verileri.get(e_alan, "")
            col_num = offset2 + e_col + 1
            _hucre_yaz(ws, row_idx, col_num, deger, veri_font, link_font, veri_alignment, ince_border)

    # ─── Sütun genişlikleri ──────────────────────────────────────────────
    for col_idx in range(1, len(tum_basliklar) + 1):
        max_len = 0
        for row_idx in range(1, len(tum_satirlar) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Filtre ve dondurma
    if tum_basliklar:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    print(f"   ✅ Excel dosyası kaydedildi: {OUTPUT_FILE}")
    return OUTPUT_FILE


def main():
    print("=" * 60)
    print("   MEBBIS Araştırma İzni - Bekleyen İşlemler")
    print("   Veri Çekme Scripti")
    print("=" * 60)

    # 1) Tarayıcıyı başlat
    driver = tarayici_baslat()

    try:
        # 2) Giriş yap
        giris_yap(driver)

        # 3) Hedef sayfaya git
        hedef_sayfaya_git(driver)

        # 4) Oturum kontrolü
        if "oturum" in driver.page_source.lower() and "bulunamadı" in driver.page_source.lower():
            print("\n⚠️ Oturum geçersiz görünüyor. Tekrar deneniyor...")
            time.sleep(2)
            driver.get(HEDEF_URL)
            time.sleep(5)

        # 5) Ana tabloyu bul
        print("\n🔍 Sayfa analiz ediliyor...")
        tablo = ana_tablo_bul(driver)
        if not tablo:
            print("\n❌ Tablo bulunamadı!")
            input("Sayfa yüklendiyse ENTER'a basın (tekrar deneyecek): ")
            time.sleep(2)
            tablo = ana_tablo_bul(driver)
            if not tablo:
                print("❌ Hâlâ tablo bulunamadı. Script sonlandırılıyor.")
                return

        # 6) Başlıkları oku
        basliklar = tablo_basliklarini_oku(tablo)
        print(f"\n   📋 Başlıklar: {basliklar}")

        # 7) Tüm sayfalardaki satırları oku
        tum_satirlar = []
        sayfa_no = 1

        while True:
            print(f"\n   📄 Sayfa {sayfa_no} okunuyor...")
            satirlar = tablo_satirlarini_oku(tablo)
            print(f"      {len(satirlar)} satır bulundu")
            tum_satirlar.extend(satirlar)

            if sayfalama_kontrol(driver):
                sayfa_no += 1
                tablo = ana_tablo_bul(driver)
                if not tablo:
                    break
            else:
                break

        print(f"\n   📊 Toplam {len(tum_satirlar)} satır veri toplandı")

        # 8) Detay sayfalarını oku
        print("\n🔎 Detay sayfaları okunuyor...")
        driver.get(HEDEF_URL)
        time.sleep(4)

        islenecek = 0
        sayfa_no = 1

        while islenecek < len(tum_satirlar):
            tablo = ana_tablo_bul(driver)
            if not tablo:
                break

            satirlar = tablo_satirlarini_oku(tablo)

            for satir_info in satirlar:
                if islenecek >= len(tum_satirlar):
                    break

                i = islenecek + 1
                detay_butonu = satir_info["detay_butonu"]

                if detay_butonu:
                    print(f"\n   [{i}/{len(tum_satirlar)}] Detay açılıyor...")
                    ana_pencere = driver.current_window_handle
                    onceki_pencereler = set(driver.window_handles)

                    try:
                        # Butona tıkla
                        try:
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", detay_butonu)
                            time.sleep(0.5)
                            detay_butonu.click()
                        except (ElementClickInterceptedException, StaleElementReferenceException):
                            driver.execute_script("arguments[0].click();", detay_butonu)

                        time.sleep(4)
                        yeni_pencereler = set(driver.window_handles) - onceki_pencereler

                        if yeni_pencereler:
                            # Yeni sekmeye geç
                            yeni_pencere = yeni_pencereler.pop()
                            driver.switch_to.window(yeni_pencere)
                            time.sleep(2)

                            detay = detay_sayfasini_oku(driver)
                            tum_satirlar[islenecek]["detay_verileri"] = detay
                            print(f"         ✅ {len(detay)} alan okundu (yeni sekme)")

                            driver.close()
                            driver.switch_to.window(ana_pencere)
                            time.sleep(1)
                        else:
                            # Aynı sayfada açıldı
                            current_url = driver.current_url
                            if current_url != HEDEF_URL:
                                detay = detay_sayfasini_oku(driver)
                                tum_satirlar[islenecek]["detay_verileri"] = detay
                                print(f"         ✅ {len(detay)} alan okundu (yeni sayfa)")
                                driver.back()
                                time.sleep(3)
                            else:
                                # Modal
                                detay = detay_sayfasini_oku(driver)
                                tum_satirlar[islenecek]["detay_verileri"] = detay
                                print(f"         ✅ {len(detay)} alan okundu (modal)")
                                try:
                                    kapat = driver.find_element(
                                        By.CSS_SELECTOR,
                                        ".modal .close, .btn-close, [data-dismiss='modal'], [data-bs-dismiss='modal']"
                                    )
                                    kapat.click()
                                    time.sleep(1)
                                except NoSuchElementException:
                                    pass

                    except Exception as e:
                        print(f"         ⚠️ Detay hatası: {e}")
                        tum_satirlar[islenecek]["detay_verileri"] = {}

                        # Güvenli geri dönüş
                        try:
                            pencereler = driver.window_handles
                            if len(pencereler) > 1:
                                for p in pencereler:
                                    if p != ana_pencere:
                                        driver.switch_to.window(p)
                                        driver.close()
                                driver.switch_to.window(ana_pencere)
                            elif driver.current_url != HEDEF_URL:
                                driver.get(HEDEF_URL)
                                time.sleep(4)
                        except Exception:
                            driver.get(HEDEF_URL)
                            time.sleep(4)
                else:
                    print(f"\n   [{i}/{len(tum_satirlar)}] Detay butonu yok, atlanıyor.")
                    tum_satirlar[islenecek]["detay_verileri"] = {}

                islenecek += 1

            # Sonraki sayfa
            if islenecek < len(tum_satirlar):
                if sayfalama_kontrol(driver):
                    sayfa_no += 1
                    time.sleep(2)
                else:
                    break

        # 9) Sonuçları yazdır
        toplam_detay = sum(1 for s in tum_satirlar if s.get("detay_verileri"))
        print("\n" + "=" * 60)
        print(f"   📊 TOPLAM SONUÇLAR")
        print(f"   Ana tablo satırları    : {len(tum_satirlar)}")
        print(f"   Ana tablo sütunları    : {len(basliklar)}")
        print(f"   Detay okunan kayıtlar  : {toplam_detay}")
        print(f"   Detay sütun sayısı     : {len(DETAY_ALANLARI)}")
        print("=" * 60)

        # 10) Excel'e aktar
        excel_dosya = excel_olustur(basliklar, tum_satirlar)

        print("\n" + "=" * 60)
        print(f"   🎉 İŞLEM TAMAMLANDI!")
        print(f"   Excel dosyası: {excel_dosya}")
        print("=" * 60)

    except KeyboardInterrupt:
        print("\n\n⚠️ İşlem kullanıcı tarafından iptal edildi.")
    except Exception as e:
        print(f"\n\n❌ Beklenmedik hata: {e}")
        traceback.print_exc()
    finally:
        print("\n🔄 Tarayıcı açık bırakılıyor. Manuel olarak kapatabilirsiniz.")


if __name__ == "__main__":
    main()
